from tkinter import *
import tkinter
import openpyxl
from ClassPerosn import Person
from tkinter import messagebox
from tkinter import ttk


class Window:
    WORK_BOOK = None
    FIND_VALUE = ''
    ERRORS_VALUE = []
    SHEET_NAME = 'Diplom Work'

    @classmethod
    def load_button(cls):
        if not cls.check_inputs():
            tkinter.messagebox.showwarning(title='Error',
                                           message=f'{Person.print_errors(cls.ERRORS_VALUE)}')
            cls.ERRORS_VALUE.clear()

        elif int(Person.check_age(text_birth.get(), text_death.get())) < 0:
            tkinter.messagebox.showwarning(title='Error',
                                           message='Input data Birth-Death non correct')
        else:
            Person.NAME.append(text_name.get())  # Insert all data in to a variables.
            Person.SURNAME.append(text_surname.get())
            Person.SEC_SURNAME.append(text_sec_surname.get())
            Person.DATE_OF_BIRTH.append(text_birth.get())
            Person.DATE_OF_DEATH.append(text_death.get())
            Person.GENDER.append(text_gender.get())
            Person.AGE.append(Person.check_age(text_birth.get(), text_death.get()))

            text_name.delete(0, END)  # Cleaning of input inputs.
            text_surname.delete(0, END)
            text_sec_surname.delete(0, END)
            text_birth.delete(0, END)
            text_death.delete(0, END)
            text_gender.delete(0, END)

    @classmethod
    def safe_file_button(cls):
        try:
            if not create_file_name.get():
                tkinter.messagebox.showwarning(title='Error.',
                                               message='Need to input a Name of File.')
            elif not create_file_name.get()[-4:] == 'xlsx':
                tkinter.messagebox.showwarning(title='Error.',
                                               message='Need to put xlsx extension.')
            else:
                sheet = cls.WORK_BOOK[cls.SHEET_NAME]  # create the new seet and safe the data.
                Person.init_table(sheet)
                Person.table_crea(sheet)
                cls.WORK_BOOK.save(create_file_name.get())  # Save new file.

                Person.NAME.clear()  # Cleaning of variable for new insert.
                Person.SURNAME.clear()
                Person.SEC_SURNAME.clear()
                Person.DATE_OF_BIRTH.clear()
                Person.DATE_OF_DEATH.clear()
                Person.GENDER.clear()
                Person.AGE.clear()
                create_file_name.delete(0, END)

        except TypeError:
            tkinter.messagebox.showwarning(title='Error',
                                           message='File not Found.Perhaps file not created.')
            pass

    @classmethod
    def create_file(cls):
        if cls.WORK_BOOK != None:
            tkinter.messagebox.showwarning(title='Error.', message='File already exist.')
        else:
            cls.WORK_BOOK = openpyxl.Workbook()  # Create new file.
            cls.WORK_BOOK.create_sheet(title=cls.SHEET_NAME, index=0)  # Create the Sheet

    @classmethod
    def load_file_button(cls):
        try:
            if not load_file_name.get():
                raise Exception
            cls.WORK_BOOK = openpyxl.load_workbook(load_file_name.get())  # Open the file.
            load_file_name.delete(0, END)
        except Exception:
            tkinter.messagebox.showwarning(title='Error', message='File not found.')

    @classmethod
    def search_button(cls):
        try:
            sheet = cls.WORK_BOOK['Diplom Work']
            rows = (sheet.max_row - 1)
            col = sheet.max_column
            se_val = search_val.get()  # search value
            array1 = sorted(set(Person.search_row(rows, 3, sheet, se_val)))
            find_value = Person.result_of_search(array1, sheet, col)
            text_uot = tkinter.Text(search_label_frame, width=50, height=10, )
            text_uot.insert(tkinter.END, Person.print_search_result(find_value, Person.age_word))
            text_uot.grid(row=40, column=1, sticky="e", padx=20, pady=10)
        except TypeError:
            tkinter.messagebox.showwarning(title='Error', message='No source file finded.')

    @classmethod
    def check_inputs(cls):
        result_check = True
        if not Person.input_str(text_name.get()) or not text_name.get():
            cls.ERRORS_VALUE.append('Error Name Input')
            result_check = False
        if not Person.input_str(text_surname.get()):
            cls.ERRORS_VALUE.append('Error Surname Input')
            result_check = False
        if not Person.input_str(text_sec_surname.get()):
            cls.ERRORS_VALUE.append('Error Second Name Input')
            result_check = False
        if not Person.check_data(text_birth.get()) or not text_birth.get():
            cls.ERRORS_VALUE.append('Error Data of Birth Input')
            result_check = False
        if text_death.get():
            date_dt = Person.check_data(text_death.get())
            if not date_dt:
                cls.ERRORS_VALUE.append('Error Data of Dead Input')
                result_check = False
        if text_gender.get():
            if text_gender.get().lower() == 'f':
                ...
            elif text_gender.get().lower() == 'm':
                ...
            else:
                cls.ERRORS_VALUE.append('Error Title Input')
                result_check = False
        else:
            cls.ERRORS_VALUE.append('Error Title Input')
            result_check = False
        return result_check

window = tkinter.Tk()
window.title("!!! DIPLOM !!!")

frame = tkinter.Frame(window)
frame.pack()

welcome_label_frame = tkinter.LabelFrame(frame, text='User Information')
welcome_label_frame.grid(row=0, column=0, sticky="news", padx=20, pady=10)


button_label_frame = tkinter.LabelFrame(frame, text='File managing Button')
button_label_frame.grid(row=14, column=0, sticky="news", padx=20, pady=10)

search_label_frame = tkinter.LabelFrame(frame, text='Search load Peoples')
search_label_frame.grid(row=22, column=0, sticky="news", padx=20, pady=10)


name_label = tkinter.Label(welcome_label_frame, text=" Put the Name: ", font=("Helvetica", 13))
name_label.grid(row=2, column=0, stick="W", padx=20, pady=10)
text_name = tkinter.Entry(welcome_label_frame, width=60)
text_name.grid(row=2, column=1, sticky="E", padx=20, pady=10)

surname_label = tkinter.Label(welcome_label_frame, text=" Put the Surname: ", font=("Helvetica", 13))
surname_label.grid(row=4, column=0, stick="W", padx=20, pady=10)
text_surname = tkinter.Entry(welcome_label_frame, width=60)
text_surname.grid(row=4, column=1, sticky="E", padx=20, pady=10)

sec_surname_label = tkinter.Label(welcome_label_frame, text=" Put the Second Surname: ", font=("Helvetica", 13))
sec_surname_label.grid(row=6, column=0, stick="W", padx=20, pady=10)
text_sec_surname = tkinter.Entry(welcome_label_frame, width=60)
text_sec_surname.grid(row=6, column=1, sticky="E", padx=20, pady=10)

birth_label = tkinter.Label(welcome_label_frame, text=" Put the Birth date: ", font=("Helvetica", 13))
birth_label.grid(row=8, column=0, stick="W", padx=20, pady=10)
text_birth = tkinter.Entry(welcome_label_frame, width=60)
text_birth.grid(row=8, column=1, sticky="E", padx=20, pady=10)

death_label = tkinter.Label(welcome_label_frame, text=" Put the Death date: ", font=("Helvetica", 13))
death_label.grid(row=10, column=0, stick="W", padx=20, pady=10)
text_death = tkinter.Entry(welcome_label_frame, width=60)
text_death.grid(row=10, column=1, sticky="E", padx=20, pady=10)

gender_label = tkinter.Label(welcome_label_frame, text=" Put the Title: ", font=("Helvetica", 13))
gender_label.grid(row=12, column=0, stick="W", padx=20, pady=10)
text_gender = ttk.Combobox(welcome_label_frame, width=57, values=['F', 'M'])
text_gender.grid(row=12, column=1, sticky="E", padx=20, pady=10)

create_file_label = tkinter.Label(button_label_frame, text=" Put the Name for create File: ", font=("Helvetica", 13))
create_file_label.grid(row=14, column=0, stick="W", padx=20, pady=10)
create_file_name = tkinter.Entry(button_label_frame, width=40)
create_file_name.grid(row=14, column=1, sticky="E", padx=20, pady=10)

load_file_label = tkinter.Label(button_label_frame, text=" Put the Name for load File: ", font=("Helvetica", 13))
load_file_label.grid(row=16, column=0, stick="W", padx=20, pady=10)
load_file_name = tkinter.Entry(button_label_frame, width=40)
load_file_name.grid(row=16, column=1, sticky="E", padx=20, pady=10)

create_button = tkinter.Button(button_label_frame, text="Create File",
                               command=Window.create_file, font=("Helvetica", 13))
create_button.grid(row=14, column=3, sticky="e", padx=10, pady=10)

load_people_button = tkinter.Button(welcome_label_frame, text="Load People",
                                    command=Window.load_button, font=("Helvetica", 13))
load_people_button.grid(row=13, column=1, sticky="news", padx=10, pady=10)

safe_button = tkinter.Button(button_label_frame, text="Safe File",
                             command=Window.safe_file_button, font=("Helvetica", 13))
safe_button.grid(row=20, column=0, sticky="news", padx=10, pady=10)
safe_button.grid_rowconfigure(1, weight=1)
safe_button.grid_columnconfigure(1, weight=1)

load_button = tkinter.Button(button_label_frame, text="Load File",
                             command=Window.load_file_button, font=("Helvetica", 13))
load_button.grid(row=20, column=1, sticky="news", padx=10, pady=10)

create_button = tkinter.Button(search_label_frame, text="Search People",
                               command=Window.search_button, font=("Helvetica", 13))
create_button.grid(row=30, column=0, )

search_label = tkinter.Label(search_label_frame, text=" Put the Search name: ", font=("Helvetica", 13))
search_label.grid(row=26, column=0, stick="W", padx=20, pady=10)
search_val = tkinter.Entry(search_label_frame, width=60)
search_val.grid(row=26, column=1, sticky="E", padx=20, pady=10)

if __name__ == "__main__":
    window.mainloop()
