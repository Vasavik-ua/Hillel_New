import openpyxl

wb = openpyxl.load_workbook('DiploM.xlsx')
sheet = wb['Diplom Work']
rows = (sheet.max_row - 1)
col = sheet.max_column
a = 'e'


def search_row(rows, col, sheet):
    array1 = []
    for item in range(col):
        for i in range(rows):
            cell = sheet.cell(row=i + 2, column=item+1)
            x = (str(cell.value.lower()).find(a.lower()))
            if x < 0:
                pass
            else:
                array1.append(i + 2)
    return array1


def result_of_search(val, sheet, col):  # Val:array;
    find_value = []
    for v in range(len(val)):
        row_list = []
        for i in range(col):
            cell = sheet.cell(row=int(list(val)[v]), column=i+1)
            row_list.append(cell.value)
        find_value.append(row_list.copy())
        row_list.clear()
    return find_value


def age_word(x):
    t = x[-1:]
    result_qty = lambda t: 'рік' if int(t) == 1 else 'роки' if 2 <= int(t) <= 4 else 'років'
    return result_qty(t)

def print_search_result(find_value):
    for i in find_value:
        final_print = []
        final_print.append(f'{i[0].title()} ')
        if not i[1] == None:
            final_print.append(f'{i[1].title()} ')
        if not i[2] == None:
            final_print.append(f'{i[2].title()} ')
        final_print.append(f'{i[6]} ')
        final_print.append(f'{age_word(str(i[6]))}, ')
        if ((i[5]).lower()) == 'm' and (not i[4] == None):
            final_print.append(f'чоловік. Народився {i[3]}. Помер {i[4]}.')
        elif ((i[5]).lower()) == 'm' and i[4] == None:
            final_print.append(f'чоловік. Народився {i[3]}.')
        if ((i[5]).lower()) == 'f' and (not i[4] == None):
            final_print.append(f'жінка. Народилася {i[3]}. Померла {i[4]}.')
        elif((i[5]).lower()) == 'f' and i[4] == None:
            final_print.append(f'жінка. Народилася {i[3]}.')
        for val in final_print:
            print(val, end='')
        print('')
        final_print.clear()


array1 = sorted(set(search_row(rows, 3, sheet)))
find_value = result_of_search(array1, sheet, col)
print_search_result(find_value)