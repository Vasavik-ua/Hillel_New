import openpyxl


class Person:
    NAME= []
    SURNAME = []
    SEC_SURNAME = []
    DATE_OF_BIRTH = []
    DATE_OF_DEATH = []
    GENDER = []
    AGE = []


    def column_name(self, shee, array):
        array.insert(0, 'NAME: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 1)
            cell.value = value
    def column_surname(self, shee, array):
        array.insert(0, 'SURNAME: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 2)
            cell.value = value
    def column_sec_surname(self, shee, array):
        array.insert(0, 'SEC SURNAME: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 3)
            cell.value = value
    def column_birth(self, shee, array):
        array.insert(0, 'DATE OF BIRTH: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 4)
            cell.value = value
    def column_death(self, shee, array):
        array.insert(0, 'DATE OF DEATH: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 5)
            cell.value = value

    def column_gender(self, shee, array):
        array.insert(0, 'GENDER: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 6)
            cell.value = value

    def column_gender(self, shee, array):
        array.insert(0, 'AGE: ')
        for row_id, value in enumerate(array):
            cell = shee.cell(row=row_id + 1, column= 7)
            cell.value = value






work_book = openpyxl.Workbook()  # Create new file.
work_book.create_sheet(title='Diplom Work', index=0)
sheet = work_book['Diplom Work']


work_book.save('DiploM.xlsx')  # Save new file.


#wb = openpyxl.load_workbook('h.w.20+.xlsx')  #  Open the file.