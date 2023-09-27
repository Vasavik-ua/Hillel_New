import openpyxl
import csv


with open('h.m.19.csv', mode='r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=',')
    new_data = []
    for item in reader:
        new_data.append(item)


def search_val(val):  # val = search element.
    for item in range(len(new_data)):
        for value in range(len(new_data[item])):
            if new_data[item][value].lower() == val.lower():
                return value


def delete_serch(val):
    for item in new_data:
        item.pop(val)
    return new_data


def write_row(sheet,col_n, rows):  # Sheet in use; witch colmn; max val of row.
    id_new = []
    for item in range(col_n, col_n+1 ):
        for val in range(1, rows + 1):
            cell = sheet.cell(row=val, column=item)
            id_new.append(cell.value)
    return id_new


def create_table(sheet, array):
    for row_id, row in enumerate(array):
        for col_id, value in enumerate(row):
            cell = sheet.cell(row=row_id + 1, column=col_id + 1)
            cell.value = value


id_search = (search_val('age'))
new_array = delete_serch(id_search)

woork_book = openpyxl.Workbook()
woork_book.create_sheet(title='First sheet', index=0)
sheet = woork_book['First sheet']



create_table(sheet, new_array)

#for row_id, row in enumerate(new_array):
#    for col_id, value in enumerate(row):
#        cell = sheet.cell(row=row_id + 1, column=col_id + 1)
#        cell.value = value



sheet_1 = woork_book['Sheet']

rows = sheet.max_row
cols = sheet.max_column


new_list = []
id_new = write_row(sheet,1,rows)
new_list.append(id_new)
name_new = write_row(sheet, 2, rows)
new_list.append(name_new)
phone_new = write_row(sheet, 3, rows)
new_list.append(phone_new)


for row_id, row in enumerate(new_list):
    for col_id, value in enumerate(row):
        cell = sheet_1.cell(row=row_id + 1, column=col_id + 1)
        cell.value = value

woork_book.save('h.w.20+.xlsx')
woork_book = openpyxl.load_workbook('h.w.20+.xlsx')
